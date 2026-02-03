#!/usr/bin/env python3
"""
Create Excel file with multiple sheets:
1. Main sheet with all data
2. Social Links sheet with just artist names and social media links
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def main():
    # Read the updated data
    print("Reading updated data...")
    df = pd.read_csv('female_singers_social_updated.csv')
    
    # Create Excel writer
    output_file = 'female_singers_final.xlsx'
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All data
        print("Creating main data sheet...")
        df.to_excel(writer, sheet_name='All Data', index=False)
        
        # Sheet 2: Social Links only
        print("Creating social links sheet...")
        social_links_df = df[[
            'Artist',
            'Artist country',
            'instagram_url',
            'instagram_handle',
            'tiktok_url',
            'tiktok_handle',
            'youtube_url',
            'youtube_channel_id',
            'soundcloud_url',
            'soundcloud_handle',
            'twitter_url',
            'twitter_handle',
            'facebook_url',
            'website_url'
        ]].copy()
        
        social_links_df.to_excel(writer, sheet_name='Social Links', index=False)
    
    # Format the Social Links sheet
    print("Formatting Social Links sheet...")
    wb = load_workbook(output_file)
    ws = wb['Social Links']
    
    # Format header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_file)
    
    print(f"\n{'='*80}")
    print("SUCCESS!")
    print(f"{'='*80}")
    print(f"✓ Created {output_file} with 2 sheets:")
    print(f"  1. 'All Data' - Complete dataset with all {len(df.columns)} columns")
    print(f"  2. 'Social Links' - Artist names and social media links only")
    print(f"\nTotal artists: {len(df)}")
    print(f"Social media fields per artist: 14")
    print(f"\nFile location:")
    print(f"  /Users/larissaivanova/Dropbox/Spotify Artist Research - Maxwell Aden/{output_file}")

if __name__ == "__main__":
    main()
