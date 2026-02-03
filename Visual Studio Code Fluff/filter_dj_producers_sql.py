hosimport pandas as pd
import sqlite3
from openpyxl import load_workbook
import os

# File paths
source_file = 'Soundcharts Pulled-Out Data/DJProducers.csv'
final_csv = 'Final_Social Links/dj_producers_final_enriched.csv'
final_xlsx = 'Final_Social Links/dj_producers_final_enriched.xlsx'

print("Loading source data from Soundcharts...")
# Read the source CSV
df_source = pd.read_csv(source_file, low_memory=False)
print(f"Source data loaded: {len(df_source)} rows")

print("\nLoading current final data...")
# Read the current final CSV
df_final = pd.read_csv(final_csv, low_memory=False)
print(f"Current final data: {len(df_final)} rows")

# Create an in-memory SQLite database
conn = sqlite3.connect(':memory:')

# Load both dataframes into SQL tables
df_source.to_sql('source', conn, index=False, if_exists='replace')
df_final.to_sql('final', conn, index=False, if_exists='replace')

print("\nData loaded into SQL database")
print("\nExecuting SQL filter query...")

# SQL query to filter artists that exist in final but get updated data from source
# This will match by Artist name and update with source data
query = """
SELECT s.*
FROM source s
INNER JOIN final f ON s.Artist = f.Artist
ORDER BY s.Artist
"""

# Execute the query
df_filtered = pd.read_sql_query(query, conn)
conn.close()

print(f"\nFiltered data: {len(df_filtered)} rows (artists that exist in both files)")

# Save the filtered data back to CSV
print(f"\nSaving filtered data to {final_csv}...")
df_filtered.to_csv(final_csv, index=False)
print("CSV file updated successfully!")

# Update the Excel file
print(f"\nUpdating Excel file {final_xlsx}...")
try:
    # Load the existing workbook
    wb = load_workbook(final_xlsx)
    
    # Get the first sheet (or you can specify sheet name)
    if 'DJ Producers' in wb.sheetnames:
        sheet_name = 'DJ Producers'
    elif 'Sheet1' in wb.sheetnames:
        sheet_name = 'Sheet1'
    else:
        sheet_name = wb.sheetnames[0]
    
    # Remove the old sheet and create a new one
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    # Create new sheet
    ws = wb.create_sheet(sheet_name, 0)
    
    # Write headers
    for col_idx, col_name in enumerate(df_filtered.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, row_data in enumerate(df_filtered.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    wb.save(final_xlsx)
    print("Excel file updated successfully!")
    
except Exception as e:
    print(f"Error updating Excel file: {e}")
    print("CSV file was updated successfully, but Excel update failed.")
    print("You may need to manually open and save the CSV as Excel.")

print("\n" + "="*60)
print("SUMMARY")
print("="*60)
print(f"Source file: {source_file} ({len(df_source)} rows)")
print(f"Final file (before): {len(df_final)} rows")
print(f"Final file (after): {len(df_filtered)} rows")
print(f"Artists filtered: {len(df_final) - len(df_filtered)} removed")
print("="*60)
print("\nFiltering complete! The final file now contains only artists")
print("that exist in both the Soundcharts source and the previous final file,")
print("with updated data from the Soundcharts source.")
