import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the CSV file
print("Reading the CSV file...")
df = pd.read_csv('rappers_final_enriched.csv')

# Define the columns we want to keep
columns_to_keep = [
    'Artist',
    'instagram_url',
    'instagram_handle',
    'tiktok_url',
    'tiktok_handle',
    'youtube_url',
    'youtube_channel_id',
    'soundcloud_url',
    'soundcloud_handle',
    'twitter_url',
    'twitter_handle',
    'facebook_url',
    'website_url'
]

# Create a new dataframe with only the selected columns
print("Extracting artist and social media columns...")
social_links_df = df[columns_to_keep].copy()

# Load the existing Excel file
print("Loading existing Excel file...")
excel_file = 'rappers_final_enriched.xlsx'
workbook = openpyxl.load_workbook(excel_file)

# Check if 'Social Links' sheet already exists and remove it
if 'Social Links' in workbook.sheetnames:
    print("Removing existing 'Social Links' sheet...")
    del workbook['Social Links']

# Create a new sheet for social links
print("Creating new 'Social Links' sheet...")
social_sheet = workbook.create_sheet('Social Links')

# Write the dataframe to the new sheet
print("Writing data to the new sheet...")
for row in dataframe_to_rows(social_links_df, index=False, header=True):
    social_sheet.append(row)

# Save the workbook
print("Saving the updated Excel file...")
workbook.save(excel_file)

print(f"\nSuccess! Added 'Social Links' sheet to {excel_file}")
print(f"Total artists: {len(social_links_df)}")
print(f"Columns included: {', '.join(columns_to_keep)}")
